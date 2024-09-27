import os
import time
from datetime import datetime
from typing import Any

import nltk
from googletrans import Translator
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from selenium.webdriver.common.by import By

from utils.chrome import Chrome as ChromeDriver

nltk.download("vader_lexicon")  # NLTK의 VADER 감정 사전 로컬 설치


class ReviewScrapper:
    def __init__(self, is_test=False) -> None:
        self.__is_test = is_test
        self.__review_list_url = "https://www.creditkarma.com/reviews/personal-loan/single/id/lending-club-personal-loans"

        self.__result_dir_path: str = ""  # # 결과 엑셀 파일이 존재하는 디렉터리
        self.__result_xl_name: str = ""  # 결과 엑셀 파일 이름
        self.__review_dict_list: list[dict[str, Any]] = []

    def run(self) -> None:
        print(f"📢 크롤링 작업 시작{': 테스트' if self.__is_test else ''}")
        self.__get_reviews()
        self.__set_result_xl()
        self.__analysis_sentiment()
        self.__open_result_xl()
        self.__write_result_xl()
        self.__save_result_xl()
        print("📢 선택하신 작업이 완료되었습니다!")

    def __set_result_xl(self) -> None:
        print("🔥 결과 엑셀 생성 중..")
        # 결과 디렉터리: /result
        self.__result_dir_path = os.path.join(os.getcwd(), "result")
        if not os.path.exists(self.__result_dir_path):
            os.mkdir(self.__result_dir_path)
        cur_time = datetime.now().strftime("%Y%m%d%H%M%S")  # 파일명에 붙일 현재 시각
        self.__result_xl_name = (
            f"reviews_{cur_time}.xlsx"  # 파일명 세팅 - "reviews_(년월일시분초)" 형태
        )

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "title"
        ws["B1"] = "contents_en"
        ws["C1"] = "contents_ko"
        ws["D1"] = "created_at"
        ws["E1"] = "rating"
        ws["F1"] = "helpful"
        ws["G1"] = "not_helpful"
        ws["H1"] = "s_compound"
        ws["I1"] = "s_positive"
        ws["J1"] = "s_neutral"
        ws["K1"] = "s_negative"

        wb.save(f"{self.__result_dir_path}/{self.__result_xl_name}")

        print(f"✅ 결과 엑셀 생성 완료: {self.__result_xl_name}")

    # @retry(work="엑셀 오픈")
    def __open_result_xl(self) -> None:
        self.workbook = load_workbook(
            f"{self.__result_dir_path}/{self.__result_xl_name}"
        )
        self.worksheet = self.workbook["Sheet"]

    # @retry(work="아이템 리스트 페이지 이동")
    def __get_reviews(self) -> None:
        with ChromeDriver() as driver:
            try:
                driver.get(self.__review_list_url)
                time.sleep(3)
            except Exception as e:
                print(f"⛔️ ERROR: {e}")
                raise Exception("-> 첫 리스트 페이지 이동 불가")

            total_review_count: int = int(
                (
                    driver.find_element(
                        By.XPATH,
                        "//span[@data-testid='reviews-total']",
                    )
                    .get_attribute("innerText")
                    .strip()
                    .split(" ")[0]
                    .strip()
                )
            )
            print(f"- 총 리뷰 개수: {total_review_count}")

        page = 0
        while True:
            page += 1
            print(f"- PAGE: {page}")
            if len(self.__review_dict_list) >= total_review_count:
                break
            if len(self.__review_dict_list) > 7360:  # 7360개 이후 에러 발생
                break
            # # TEST
            # if page >= 700: # 페이지 범위: 이하
            #     break
            # if page < 650: # 페이지 범위: 초과
            #     continue
            try:
                with ChromeDriver() as driver:
                    driver.get(f"{self.__review_list_url}?pg={page}")
                    time.sleep(3)

                    # More 버튼 클릭
                    more_btn_list: list = driver.find_elements(
                        By.XPATH,
                        "//span[text()='More']",
                    )
                    for more_btn in more_btn_list:
                        more_btn.click()
                        time.sleep(1)
                    print("💡 More 버튼 모두 클릭 완료")

                    # Review 리스트 추출
                    review_div_list: list = driver.find_elements(
                        By.XPATH,
                        "//section[@id='top-of-reviews']/div/div",
                    )
                    # print('review_div_list', len(review_div_list))
                    for review_div in review_div_list:
                        print(
                            f"👉 {len(self.__review_dict_list)+1}/{total_review_count}"
                        )
                        review_dict = {}

                        # Title
                        title = review_div.find_element(
                            By.XPATH, ".//h5[contains(@class, 'lh-title')]"
                        ).text.strip()
                        review_dict["title"] = title
                        # print(f'- title: {title}')

                        # Contents
                        contents: str = review_div.find_element(
                            By.XPATH, ".//p[contains(@class, 'lh-copy')]"
                        ).text.strip()
                        review_dict["contents"] = contents
                        # print(f'- contents: {contents}')

                        # Created At: MMM DD, YYYY
                        created_at_str: str = review_div.find_element(
                            By.XPATH,
                            ".//span[contains(@class, 'self-start')]",
                        ).text.strip()
                        create_at_obj = datetime.strptime(created_at_str, "%b %d, %Y")
                        created_at = create_at_obj.strftime("%Y-%m-%d")
                        review_dict["created_at"] = created_at
                        # print(f'- created_at: {created_at}')

                        # Rating
                        rating: float = float(
                            review_div.find_element(By.XPATH, ".//div[@role='img']")
                            .get_attribute("aria-label")
                            .strip()
                            .replace("Rating:", "")
                            .strip()
                            .split(" ")[0]
                        )
                        review_dict["rating"] = rating
                        # print(f'- rating: {rating}')

                        # Helpful / Not Helpful
                        [helpful, not_helpful] = [
                            int(el.text.strip())
                            for el in review_div.find_elements(
                                By.XPATH, ".//span[@data-testid='action-total']"
                            )
                        ]
                        review_dict["helpful"] = helpful
                        review_dict["not_helpful"] = not_helpful
                        # print(f'- helpful: {helpful}')
                        # print(f'- not_helpful: {not_helpful}')

                        self.__review_dict_list.append(review_dict)
                        print(f"- review_dict: {review_dict}")
            except Exception as e:
                print(f"⛔️ ERROR: {e}")
                break

            # for TEST
            if self.__is_test:
                break

    def __analysis_sentiment(self):
        sia = SentimentIntensityAnalyzer()
        for idx, review_dict in enumerate(self.__review_dict_list):
            print(f"🔥 감정 분석 중.. {idx+1}/{len(self.__review_dict_list)}")
            sentiment_dict = sia.polarity_scores(review_dict["contents"])
            review_dict["s_compound"] = sentiment_dict["compound"]
            review_dict["s_positive"] = sentiment_dict["pos"]
            review_dict["s_neutral"] = sentiment_dict["neu"]
            review_dict["s_negative"] = sentiment_dict["neg"]
            time.sleep(0.1)
        print("✅ 감정 분석 완료")

    def __write_result_xl(self):
        for row_idx, review_dict in enumerate(self.__review_dict_list):
            print(f"🔥 엑셀 입력 중.. {row_idx+1}/{len(self.__review_dict_list)}")
            # title
            cell = self.worksheet[f"A{row_idx+2}"]
            cell.value = review_dict["title"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # contents_en
            cell = self.worksheet[f"B{row_idx+2}"]
            cell.value = review_dict["contents"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # contents_ko
            cell = self.worksheet[f"C{row_idx+2}"]
            translated_text = (
                Translator()
                .translate(review_dict["contents"], src="en", dest="ko")
                .text
            )
            cell.value = translated_text
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # created_at
            cell = self.worksheet[f"D{row_idx+2}"]
            cell.value = review_dict["created_at"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # rating
            cell = self.worksheet[f"E{row_idx+2}"]
            cell.value = review_dict["rating"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # helpful
            cell = self.worksheet[f"F{row_idx+2}"]
            cell.value = review_dict["helpful"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # not_helpful
            cell = self.worksheet[f"G{row_idx+2}"]
            cell.value = review_dict["not_helpful"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # compound
            cell = self.worksheet[f"H{row_idx+2}"]
            cell.value = review_dict["s_compound"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # positive
            cell = self.worksheet[f"I{row_idx+2}"]
            cell.value = review_dict["s_positive"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # neutral
            cell = self.worksheet[f"J{row_idx+2}"]
            cell.value = review_dict["s_neutral"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
            # negative
            cell = self.worksheet[f"K{row_idx+2}"]
            cell.value = review_dict["s_negative"]
            cell.alignment = Alignment(wrapText=False, vertical="top")
        print("✅ 엑셀 입력 완료")

    # @retry(work="엑셀 저장")
    def __save_result_xl(self):
        print("🔥 엑셀 저장 중..")
        if os.path.isfile(f"{self.__result_dir_path}/{self.__result_xl_name}"):
            self.workbook.close()
        self.workbook.save(f"{self.__result_dir_path}/{self.__result_xl_name}")
        time.sleep(2)
        print(f"🔎 엑셀 파일명: {self.__result_xl_name}")
        time.sleep(1)
        print("✅ 엑셀 저장 완료")
